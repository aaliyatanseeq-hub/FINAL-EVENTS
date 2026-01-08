"""
ULTRA-STRICT Event Intelligence Platform
WITH POSTGRESQL DATABASE LAYER
FIXED: Uses OAuth 1.1 for all Twitter actions (comments, retweets, likes)
"""

from fastapi import FastAPI, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.types import ASGIApp
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from contextlib import asynccontextmanager
import uvicorn
import re
import time
import os
from pathlib import Path
from sqlalchemy.orm import Session
from sqlalchemy import func, text
from datetime import datetime
from engines.event_engine import SmartEventEngine
from engines.attendee_engine import SmartAttendeeEngine
from services.twitter_client import TwitterClient
from services.oauth_twitter_client import OAuthTwitterClient
from database import crud, schemas, get_db, init_database
from database.crud import generate_cache_key

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    init_database()
    print("✅ PostgreSQL database initialized")
    yield
    # Shutdown (if needed)
    pass

app = FastAPI(
    title="Event Intelligence Platform",
    description="FIXED: Uses OAuth 1.1 for all Twitter actions + PostgreSQL Database",
    version="2.1.0",
    lifespan=lifespan
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:4000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Middleware to disable caching for static files during development
class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        # Add no-cache headers for static files
        if request.url.path.startswith("/static/"):
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
        return response

app.add_middleware(NoCacheMiddleware)

# Initialize engines
event_engine = SmartEventEngine()
attendee_engine = SmartAttendeeEngine()

class EventDiscoveryRequest(BaseModel):
    location: str
    start_date: str
    end_date: str
    categories: List[str]
    max_results: int

class AttendeeDiscoveryRequest(BaseModel):
    event_name: str
    event_date: Optional[str] = None
    max_results: int

class TwitterActionRequest(BaseModel):
    attendees: List[dict]
    message: Optional[str] = None

class CommentRequest(BaseModel):
    posts: List[dict]
    comment_template: Optional[str] = None
    hashtags: Optional[str] = None

@app.get("/api/health")
async def health_check(db: Session = Depends(get_db)):
    twitter_client = TwitterClient()
    
    # Check database connection
    try:
        db.execute(text("SELECT 1"))
        db_status = "healthy"
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    # Debug: Check crud module
    try:
        crud_info = {
            "crud_type": str(type(crud)) if crud else "None",
            "crud_is_none": crud is None,
            "has_update_analytics": hasattr(crud, 'update_analytics') if crud else False,
            "crud_dir_count": len([x for x in dir(crud) if not x.startswith('_')]) if crud else 0,
            "update_analytics_exists": 'update_analytics' in dir(crud) if crud else False
        }
    except Exception as e:
        crud_info = {
            "error": str(e),
            "crud_type": "error",
            "crud_is_none": True
        }
    
    return {
        "status": "healthy",
        "database": db_status,
        "twitter_search_ready": twitter_client.is_operational(),
        "twitter_actions_ready": twitter_client.api_v1 is not None,
        "features": ["event_discovery", "attendee_discovery", "twitter_actions", "database"],
        "debug_crud": crud_info
    }

@app.get("/api/auth-status")
async def auth_status():
    """Check which authentication methods are working"""
    from services.twitter_client import TwitterClient
    
    twitter_client = TwitterClient()
    
    # Test OAuth 1.1
    oauth1_working = False
    oauth1_user = None
    if twitter_client.api_v1:
        try:
            user = twitter_client.api_v1.verify_credentials()
            oauth1_working = True
            oauth1_user = user.screen_name
        except Exception as e:
            print(f"OAuth 1.1 test failed: {e}")
    
    return {
        "oauth1_ready": oauth1_working,
        "oauth1_user": oauth1_user,
        "recommendation": "Using OAuth 1.1 for all actions"
    }

@app.post("/api/discover-events")
async def discover_events(
    request: EventDiscoveryRequest, 
    db: Session = Depends(get_db)
):
    """STRICT: Only called when user explicitly requests events"""
    try:
        print(f"🎯 EVENT REQUEST: {request.max_results} events in {request.location}")
        
        # Import location validator
        from services.location_validator import LocationValidator
        validator = LocationValidator()
        
        # ============================================
        # STEP 1: VALIDATE LOCATION
        # ============================================
        print(f"🔍 Validating location: '{request.location}'")
        is_valid_location, normalized_location, location_error = validator.validate_location(request.location)
        
        if not is_valid_location:
            error_msg = location_error or f"'{request.location}' is not a valid location. Please enter a city name (e.g., 'New York', 'London', 'Los Angeles')"
            print(f"❌ Location validation failed: {error_msg}")
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "Invalid location",
                    "message": error_msg,
                    "location": request.location,
                    "suggestion": "Please enter a valid city name (e.g., 'New York', 'London', 'Los Angeles')"
                }
            )
        
        print(f"✅ Location validated: '{request.location}' → '{normalized_location}'")
        # Use normalized location (Pydantic models are immutable, so we'll use normalized_location in the request)
        if normalized_location:
            # Update the request object with normalized location
            request = request.model_copy(update={"location": normalized_location})
        
        # ============================================
        # STEP 2: VALIDATE DATE RANGE
        # ============================================
        print(f"🔍 Validating date range: {request.start_date} to {request.end_date}")
        is_valid_dates, date_error, start_dt, end_dt = validator.validate_date_range(
            request.start_date, 
            request.end_date
        )
        
        if not is_valid_dates:
            print(f"❌ Date validation failed: {date_error}")
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "Invalid date range",
                    "message": date_error,
                    "start_date": request.start_date,
                    "end_date": request.end_date
                }
            )
        
        if start_dt is not None and end_dt is not None:
            print(f"✅ Date range validated: {start_dt.date()} to {end_dt.date()} ({((end_dt - start_dt).days)} days)")
        else:
            print(f"✅ Date range validated: {request.start_date} to {request.end_date}")
        
        # Debug: Check crud module
        print(f"DEBUG: crud module type: {type(crud)}")
        print(f"DEBUG: crud has update_analytics: {hasattr(crud, 'update_analytics')}")
        if crud is None:
            raise HTTPException(status_code=500, detail="crud module is None!")
        if not hasattr(crud, 'update_analytics'):
            available = [x for x in dir(crud) if not x.startswith('_')]
            raise HTTPException(status_code=500, detail=f"update_analytics not found in crud. Available: {available}")
        
        # Update analytics
        try:
            crud.update_analytics(db, "event_search")
            print("✅ Analytics updated successfully")
        except Exception as e:
            print(f"⚠️ Analytics update failed: {e}")
            import traceback
            traceback.print_exc()
            # Don't fail the request, just log the error
        
        # Check cache first
        try:
            cache_key = generate_cache_key("event_discovery", request.model_dump())
            cached_result = crud.get_cache(db, cache_key)
            
            if cached_result is not None:
                print("✅ Returning cached events")
                try:
                    crud.update_analytics(db, "api_saved")
                except:
                    pass
                return {
                    "success": True,
                    "events": cached_result.get("events", []),
                    "total_events": cached_result.get("total_events", 0),
                    "requested_limit": request.max_results,
                    "source": "cache"
                }
        except Exception as e:
            print(f"⚠️ Cache check failed: {e}")
        
        # Validate limits
        if request.max_results > 100:
            request.max_results = 100
        if request.max_results < 1:
            request.max_results = 1

        # ============================================
        # STEP 3: SMART CACHING - Check for overlapping date ranges
        # ============================================
        cached_events = []
        cached_event_dicts = []
        try:
            # Get cached events that fall within the requested date range
            # This checks for ANY events in the database that overlap with the requested range
            cached_events = crud.get_cached_events_by_date_range(
                db=db,
                location=request.location,
                start_date=request.start_date,
                end_date=request.end_date,
                categories=request.categories
            )
            
            if cached_events:
                print(f"✅ Found {len(cached_events)} cached events in database for date range {request.start_date} to {request.end_date}")
                # Convert cached events to dicts
                for event in cached_events:
                    try:
                        # Use manual conversion to avoid schema issues
                        # Generate event_hash for deduplication (same as engine)
                        import hashlib
                        import re
                        normalized = f"{event.event_name.lower()}_{(event.exact_venue or '').lower()}_{(event.exact_date or '')[:10]}"
                        normalized = re.sub(r'[^\w\s]', '', normalized)
                        normalized = re.sub(r'\s+', '_', normalized)
                        event_hash = hashlib.md5(normalized.encode()).hexdigest()[:16]
                        
                        event_dict = {
                            "event_name": event.event_name,
                            "exact_date": event.exact_date,
                            "exact_venue": event.exact_venue or "",
                            "location": event.location,
                            "category": event.category or "other",
                            "source_url": event.source_url or "",
                            "posted_by": event.posted_by or "Unknown",
                            "hype_score": float(getattr(event, 'hype_score', 0.0) or 0.0),
                            "source": getattr(event, 'source', 'database'),
                            "confidence_score": float(getattr(event, 'confidence_score', 0.0) or 0.0),
                            "event_hash": event_hash  # Add hash for deduplication
                        }
                        cached_event_dicts.append(event_dict)
                    except Exception as e:
                        print(f"⚠️ Error converting cached event {getattr(event, 'id', 'unknown')}: {e}")
                        continue
                
                # Sort cached events by hype_score (descending)
                cached_event_dicts.sort(key=lambda x: x.get('hype_score', 0.0), reverse=True)
                print(f"📊 Cached events ready: {len(cached_event_dicts)} events sorted by hype score")
        except Exception as e:
            print(f"⚠️ Error getting cached events: {e}")
            import traceback
            traceback.print_exc()
        
        # Calculate how many more events we need
        events_needed = max(0, request.max_results - len(cached_event_dicts))
        
        if events_needed == 0:
            # We have enough cached events!
            # But we still need to enforce 50/25/25 ratio
            print(f"✅ Using {len(cached_event_dicts)} cached events (no API calls needed)")
            
            # Enforce 50/25/25 ratio on cached events
            events_by_source = {'serpapi': [], 'predicthq': [], 'ticketmaster': []}
            other_events = []
            
            for event in cached_event_dicts:
                source = event.get('source', '').lower()
                if source in events_by_source:
                    events_by_source[source].append(event)
                else:
                    other_events.append(event)
            
            # Calculate target counts for 50/25/25 ratio
            serpapi_target = int(request.max_results * 0.5)  # 50%
            predicthq_target = int(request.max_results * 0.25)  # 25%
            ticketmaster_target = int(request.max_results * 0.25)  # 25%
            
            # Adjust if rounding causes total to be less than max_results
            total_target = serpapi_target + predicthq_target + ticketmaster_target
            if total_target < request.max_results:
                serpapi_target += (request.max_results - total_target)
            
            # Sort each source by hype_score
            for source in events_by_source:
                events_by_source[source].sort(key=lambda x: x.get('hype_score', 0.0), reverse=True)
            
            # Build final list respecting 50/25/25 ratio
            ratio_filtered_events = []
            ratio_filtered_events.extend(events_by_source['serpapi'][:serpapi_target])
            ratio_filtered_events.extend(events_by_source['predicthq'][:predicthq_target])
            ratio_filtered_events.extend(events_by_source['ticketmaster'][:ticketmaster_target])
            
            # Fill remaining slots if we don't have enough from all sources
            remaining_slots = request.max_results - len(ratio_filtered_events)
            if remaining_slots > 0:
                # Fill from other sources first
                ratio_filtered_events.extend(other_events[:remaining_slots])
                remaining_slots = request.max_results - len(ratio_filtered_events)
                
                # Then fill from sources that have more events available
                if remaining_slots > 0:
                    if len(events_by_source['serpapi']) > serpapi_target:
                        ratio_filtered_events.extend(events_by_source['serpapi'][serpapi_target:serpapi_target + remaining_slots])
                    remaining_slots = request.max_results - len(ratio_filtered_events)
                    
                    if remaining_slots > 0 and len(events_by_source['predicthq']) > predicthq_target:
                        ratio_filtered_events.extend(events_by_source['predicthq'][predicthq_target:predicthq_target + remaining_slots])
                    remaining_slots = request.max_results - len(ratio_filtered_events)
                    
                    if remaining_slots > 0 and len(events_by_source['ticketmaster']) > ticketmaster_target:
                        ratio_filtered_events.extend(events_by_source['ticketmaster'][ticketmaster_target:ticketmaster_target + remaining_slots])
            
            # Limit and sort
            ratio_filtered_events = ratio_filtered_events[:request.max_results]
            ratio_filtered_events.sort(key=lambda x: x.get('hype_score', 0.0), reverse=True)
            
            try:
                crud.update_analytics(db, "api_saved")
            except:
                pass
            
            return {
                "success": True,
                "events": ratio_filtered_events,
                "total_events": len(ratio_filtered_events),
                "requested_limit": request.max_results,
                "source": "cache",
                "cached_count": len(cached_event_dicts),
                "api_calls_saved": True
            }
        
        # ============================================
        # STEP 4: FETCH MISSING EVENTS FROM APIs
        # ============================================
        print(f"🔍 Need {events_needed} more events, fetching from APIs...")
        print(f"📊 Using {len(cached_event_dicts)} cached events + fetching {events_needed} new events")
        
        try:
            # Fetch only the missing count from APIs
            engine_events = event_engine.discover_events(
                location=request.location,
                start_date=request.start_date,
                end_date=request.end_date,
                categories=request.categories,
                max_results=events_needed  # Only fetch what we need
            )
        except ValueError as e:
            # Validation errors from event engine (date range, etc.)
            print(f"❌ Event engine validation error: {e}")
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "Validation error",
                    "message": str(e)
                }
            )
        except Exception as e:
            print(f"❌ Event engine failed: {e}")
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Event discovery failed: {str(e)}")
        
        # Import quality filter
        from services.event_quality_filter import EventQualityFilter
        quality_filter = EventQualityFilter()
        
        # Store events in database with quality filtering
        events_data = []
        filtered_count = 0
        for event in engine_events:
            try:
                # Convert dataclass to dict, handling datetime objects
                # Exclude fields not in Event model (but keep event_hash for deduplication)
                event_dict = {}
                excluded_fields = {'start_datetime', 'ticket_url', 'price_range'}  # Keep event_hash for deduplication
                
                for key, value in event.__dict__.items():
                    if key in excluded_fields:
                        continue  # Skip fields not in Event model
                    if isinstance(value, datetime):
                        event_dict[key] = value.isoformat()
                    elif value is None:
                        event_dict[key] = None
                    elif key == 'location' and isinstance(value, list):
                        # CRITICAL FIX: Convert location list to string
                        # Bug fix: Empty list [] should not become '[]' string
                        if value:
                            # Non-empty list: join valid values, fallback to original location if empty result
                            joined = ', '.join(str(v) for v in value if v)
                            event_dict[key] = joined if joined else request.location
                        else:
                            # Empty list: fall back to original request location
                            event_dict[key] = request.location
                    else:
                        event_dict[key] = value
                
                # Ensure source is included
                if 'source' not in event_dict:
                    event_dict['source'] = event.source if hasattr(event, 'source') else 'unknown'
                
                # Store ticket_url and price_range in source_data if available
                source_data = event_dict.copy()
                if hasattr(event, 'ticket_url') and event.ticket_url:
                    source_data['ticket_url'] = event.ticket_url
                if hasattr(event, 'price_range') and event.price_range:
                    source_data['price_range'] = event.price_range
                
                event_dict['source_data'] = source_data
                
                # === ENTERPRISE QUALITY FILTERING ===
                # Apply quality filter to remove noise
                quality_result = quality_filter.filter_event(event_dict, request.location)
                
                # Skip events marked as noise (unless quality score is borderline)
                if not quality_result.is_real_event:
                    filtered_count += 1
                    if filtered_count <= 5:  # Log first 5 filtered events
                        print(f"      🚫 Filtered noise: {event_dict.get('event_name', 'Unknown')[:50]} - Reasons: {', '.join(quality_result.rejection_reasons[:2])}")
                    continue  # Skip storing noise events
                
                # Update category with clean category if different
                if quality_result.clean_category and quality_result.clean_category != event_dict.get('category', '').lower():
                    event_dict['category'] = quality_result.clean_category
                
                # CRITICAL FIX: Check if quality columns exist BEFORE adding them
                # This prevents database errors if migration hasn't been run
                from database import models
                valid_columns = {col.name for col in models.Event.__table__.columns}
                quality_fields = {'is_real_event', 'quality_score', 'clean_category', 'rejection_reasons', 'quality_confidence'}
                
                # Only add quality fields if columns exist in database
                for field in quality_fields:
                    if field in valid_columns:
                        if field == 'is_real_event':
                            event_dict[field] = quality_result.is_real_event
                        elif field == 'quality_score':
                            event_dict[field] = quality_result.quality_score
                        elif field == 'clean_category':
                            event_dict[field] = quality_result.clean_category
                        elif field == 'rejection_reasons':
                            event_dict[field] = quality_result.rejection_reasons
                        elif field == 'quality_confidence':
                            event_dict[field] = quality_result.confidence
                    # If column doesn't exist, don't add it (prevents DB errors)
                
                events_data.append(event_dict)
                
                # Store in database with proper error handling
                try:
                    crud.create_event(db, event_dict)
                except Exception as e:
                    print(f"⚠️ Failed to store event in DB: {e}")
                    # Don't fail the entire request, just log and continue
            except Exception as e:
                print(f"⚠️ Failed to process event: {e}")
        
        # ============================================
        # STEP 5: COMBINE CACHED + NEW EVENTS & ENFORCE 50/25/25 RATIO
        # ============================================
        # Combine cached events with newly fetched events
        all_events_combined = cached_event_dicts + events_data
        
        # Remove duplicates based on event name, date, and location
        # Use hash-based deduplication (same as engine) for consistency
        seen_event_keys = set()
        seen_hashes = set()  # Also track hashes for better deduplication
        unique_events = []
        for event in all_events_combined:
            # Create unique key for deduplication (name + date + location)
            event_key = (
                event.get('event_name', '').lower().strip(),
                event.get('exact_date', '').strip(),
                event.get('location', '').lower().strip()
            )
            
            # Also check hash if available (from engine deduplication)
            event_hash = event.get('event_hash') or None
            
            # Skip if duplicate (by key or hash)
            is_duplicate = False
            if event_key in seen_event_keys:
                is_duplicate = True
            elif event_hash and event_hash in seen_hashes:
                is_duplicate = True
            
            if not is_duplicate and event_key[0]:  # Skip empty names
                seen_event_keys.add(event_key)
                if event_hash:
                    seen_hashes.add(event_hash)
                unique_events.append(event)
        
        # ============================================
        # STEP 6: ENFORCE 50/25/25 RATIO ON FINAL EVENTS
        # ============================================
        # Group events by source to enforce ratio
        events_by_source = {'serpapi': [], 'predicthq': [], 'ticketmaster': []}
        other_events = []
        
        for event in unique_events:
            source = event.get('source', '').lower()
            if source in events_by_source:
                events_by_source[source].append(event)
            else:
                other_events.append(event)
        
        # Calculate target counts for 50/25/25 ratio
        serpapi_target = int(request.max_results * 0.5)  # 50%
        predicthq_target = int(request.max_results * 0.25)  # 25%
        ticketmaster_target = int(request.max_results * 0.25)  # 25%
        
        # Adjust if rounding causes total to be less than max_results
        total_target = serpapi_target + predicthq_target + ticketmaster_target
        if total_target < request.max_results:
            serpapi_target += (request.max_results - total_target)
        
        # Sort each source by hype_score
        for source in events_by_source:
            events_by_source[source].sort(key=lambda x: x.get('hype_score', 0.0), reverse=True)
        
        # Build final list respecting 50/25/25 ratio
        final_events = []
        final_events.extend(events_by_source['serpapi'][:serpapi_target])
        final_events.extend(events_by_source['predicthq'][:predicthq_target])
        final_events.extend(events_by_source['ticketmaster'][:ticketmaster_target])
        
        # Fill remaining slots if we don't have enough from all sources
        remaining_slots = request.max_results - len(final_events)
        if remaining_slots > 0:
            # Fill from other sources first
            final_events.extend(other_events[:remaining_slots])
            remaining_slots = request.max_results - len(final_events)
            
            # Then fill from sources that have more events available
            if remaining_slots > 0:
                # Prioritize sources that haven't met their target
                if len(events_by_source['serpapi']) > serpapi_target:
                    final_events.extend(events_by_source['serpapi'][serpapi_target:serpapi_target + remaining_slots])
                remaining_slots = request.max_results - len(final_events)
                
                if remaining_slots > 0 and len(events_by_source['predicthq']) > predicthq_target:
                    final_events.extend(events_by_source['predicthq'][predicthq_target:predicthq_target + remaining_slots])
                remaining_slots = request.max_results - len(final_events)
                
                if remaining_slots > 0 and len(events_by_source['ticketmaster']) > ticketmaster_target:
                    final_events.extend(events_by_source['ticketmaster'][ticketmaster_target:ticketmaster_target + remaining_slots])
        
        # Limit to max_results and re-sort by hype_score
        final_events = final_events[:request.max_results]
        final_events.sort(key=lambda x: x.get('hype_score', 0.0), reverse=True)
        
        print(f"📊 Final result: {len(cached_event_dicts)} cached + {len(events_data)} new = {len(unique_events)} total (returning {len(final_events)})")
        
        # Cache the combined result
        try:
            cache_data = {
                "events": final_events,
                "total_events": len(final_events),
                "cached_count": len(cached_event_dicts),
                "new_count": len(events_data)
            }
            # Cache for 3 months (90 days = 129,600 minutes)
            crud.set_cache(db, cache_key, cache_data, ttl_minutes=129600)
        except Exception as e:
            print(f"⚠️ Cache storage failed: {e}")

        return {
            "success": True,
            "events": final_events,
            "total_events": len(final_events),
            "requested_limit": request.max_results,
            "source": "combined",
            "cached_count": len(cached_event_dicts),
            "new_count": len(events_data),
            "api_calls_saved": len(cached_event_dicts) > 0
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ ERROR in discover_events: {e}")
        print(error_trace)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/api/discover-attendees")
async def discover_attendees(
    request: AttendeeDiscoveryRequest, 
    db: Session = Depends(get_db)
):
    """STRICT: Only called when user explicitly requests attendees"""
    try:
        print(f"🎯 ATTENDEE REQUEST: {request.max_results} attendees for {request.event_name}")
        
        # Update analytics
        try:
            crud.update_analytics(db, "attendee_search")
        except Exception as e:
            print(f"⚠️ Analytics update failed: {e}")
        
        # Check cache first
        try:
            cache_key = generate_cache_key("attendee_discovery", request.model_dump())
            cached_result = crud.get_cache(db, cache_key)
            
            if cached_result is not None:
                print("✅ Returning cached attendees")
                try:
                    crud.update_analytics(db, "api_saved")
                except:
                    pass
                return {
                    "success": True,
                    "attendees": cached_result.get("attendees", []),
                    "total_attendees": cached_result.get("total_attendees", 0),
                    "requested_limit": request.max_results,
                    "source": "cache"
                }
        except Exception as e:
            print(f"⚠️ Cache check failed: {e}")
        
        # Validate limits
        if request.max_results > 100:
            request.max_results = 100
        if request.max_results < 1:
            request.max_results = 1

        # Check database for existing attendees
        try:
            db_attendees = crud.get_attendees_by_event(
                db=db,
                event_name=request.event_name,
                limit=request.max_results
            )
            
            if len(db_attendees) >= request.max_results * 0.5:  # If we have at least 50% in DB
                print(f"✅ Found {len(db_attendees)} attendees in database")
                attendees = []
                for attendee in db_attendees:
                    try:
                        attendee_dict = schemas.Attendee.model_validate(attendee).model_dump()
                        attendees.append(attendee_dict)
                    except Exception as e:
                        print(f"⚠️ Schema conversion failed for attendee {attendee.id}: {e}")
                        # Fallback: manual conversion
                        # Ensure location is always a string
                        location_value = attendee.location if hasattr(attendee, 'location') else ""
                        if location_value is None:
                            location_value = ""
                        elif not isinstance(location_value, str):
                            if isinstance(location_value, (int, float)):
                                location_value = ""  # Don't show numbers as location
                            else:
                                location_value = str(location_value).strip()
                        else:
                            location_value = str(location_value).strip()
                        
                        attendees.append({
                            "username": attendee.username,
                            "display_name": attendee.display_name,
                            "bio": attendee.bio,
                            "location": location_value,  # Always a string
                            "followers_count": int(getattr(attendee, 'followers_count', 0) or 0),
                            "verified": attendee.verified or False,
                            "engagement_type": attendee.engagement_type,
                            "post_content": attendee.post_content,
                            "post_date": attendee.post_date,
                            "post_link": attendee.post_link,
                            "relevance_score": attendee.relevance_score or 0.0,
                            "event_name": attendee.event_name,
                            "source": attendee.source if hasattr(attendee, 'source') else 'unknown',
                            "user_id": getattr(attendee, 'user_id', None)  # Twitter user ID for DMs
                        })
                
                # Cache the result
                try:
                    cache_data = {
                        "attendees": attendees,
                        "total_attendees": len(attendees)
                    }
                    # Cache for 3 months (90 days = 129,600 minutes)
                    crud.set_cache(db, cache_key, cache_data, ttl_minutes=129600)
                except:
                    pass
                
                return {
                    "success": True,
                    "attendees": attendees[:request.max_results],
                    "total_attendees": len(attendees),
                    "requested_limit": request.max_results,
                    "source": "database"
                }
        except Exception as e:
            print(f"⚠️ Database query failed: {e}")
        
        # If not enough in DB, use engine
        print("🔍 Not enough in DB, using Twitter API...")
        try:
            engine_attendees = attendee_engine.discover_attendees(
            event_name=request.event_name,
            event_date=request.event_date,
            max_results=request.max_results
        )
        except Exception as e:
            print(f"❌ Attendee engine failed: {e}")
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Attendee discovery failed: {str(e)}")
        
        # Store attendees in database
        attendees_data = []
        for attendee in engine_attendees:
            try:
                # Convert dataclass to dict
                attendee_dict = {}
                for key, value in attendee.__dict__.items():
                    if isinstance(value, datetime):
                        attendee_dict[key] = value.isoformat()
                    elif value is None:
                        attendee_dict[key] = None
                    else:
                        attendee_dict[key] = value
                
                # Ensure location is always a string (never None, never a number)
                if 'location' in attendee_dict:
                    loc_value = attendee_dict['location']
                    if loc_value is None:
                        attendee_dict['location'] = ""
                    elif not isinstance(loc_value, str):
                        # If it's a number or other type, convert to string or empty
                        if isinstance(loc_value, (int, float)):
                            attendee_dict['location'] = ""  # Don't show numbers as location
                        else:
                            attendee_dict['location'] = str(loc_value).strip()
                    else:
                        attendee_dict['location'] = str(loc_value).strip()
                else:
                    attendee_dict['location'] = ""  # Default to empty string if missing
                
                # Ensure followers_count is an integer
                if 'followers_count' in attendee_dict:
                    try:
                        attendee_dict['followers_count'] = int(attendee_dict['followers_count']) if attendee_dict['followers_count'] else 0
                    except (ValueError, TypeError):
                        attendee_dict['followers_count'] = 0
                
                # Ensure source is included
                if 'source' not in attendee_dict:
                    attendee_dict['source'] = attendee.source if hasattr(attendee, 'source') else 'unknown'
                
                attendee_dict['source_data'] = attendee_dict.copy()  # Store raw data
                attendee_dict['event_name'] = request.event_name  # Ensure event_name is set
                attendees_data.append(attendee_dict)
                
                # Store in database with proper transaction handling
                try:
                    crud.create_attendee(db, attendee_dict)
                except Exception as e:
                    print(f"⚠️ Failed to store attendee in DB: {e}")
                    # Rollback the failed transaction to allow subsequent saves
                    try:
                        db.rollback()
                    except:
                        pass
                    # Don't fail the entire request, just log and continue
            except Exception as e:
                print(f"⚠️ Failed to process attendee: {e}")
        
        # Cache the result
        try:
            cache_data = {
                "attendees": attendees_data,
                "total_attendees": len(attendees_data)
            }
            crud.set_cache(db, cache_key, cache_data, ttl_minutes=30)
        except Exception as e:
            print(f"⚠️ Cache storage failed: {e}")

        return {
            "success": True,
            "attendees": attendees_data,
            "total_attendees": len(attendees_data),
            "requested_limit": request.max_results,
            "source": "api"
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ ERROR in discover_attendees: {e}")
        print(error_trace)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/api/retweet-posts")
async def retweet_posts(
    request: TwitterActionRequest, 
    db: Session = Depends(get_db)
):
    """FIXED: Retweet posts using v2 API with database logging"""
    try:
        print(f"🔄 RETWEETING {len(request.attendees)} posts")
        
        # Update analytics
        crud.update_analytics(db, "twitter_action")
        
        twitter_client = TwitterClient()
        
        if not twitter_client.is_operational():
            return {"success": False, "error": "Twitter client not operational"}

        results = []
        successful_retweets = 0
        
        for attendee in request.attendees:
            try:
                username = attendee.get('username', '')
                post_link = attendee.get('post_link', '')
                
                if not post_link:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'No post link available'
                    })
                    continue
                
                tweet_id = extract_tweet_id(post_link)
                if not tweet_id:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'Could not extract tweet ID from link'
                    })
                    continue
                
                print(f"   🔄 Retweeting {username}'s tweet: {tweet_id}")
                
                # FIXED: Use client_v2.retweet_tweet
                retweet_result = twitter_client.retweet_tweet(tweet_id)
                
                # Log action to database
                action_data = schemas.UserActionCreate(
                    action_type="retweet",
                    target_username=username,
                    target_tweet_id=tweet_id,
                    status="success" if retweet_result else "failed",
                    message=f"Retweeted post from {username}",
                    additional_data={"post_link": post_link}
                )
                crud.create_user_action(db, action_data)
                
                if retweet_result:
                    successful_retweets += 1
                    results.append({
                        'username': username,
                        'status': 'retweeted',
                        'tweet_id': tweet_id,
                        'message': f'Successfully retweeted post from {username}'
                    })
                    print(f"   ✅ Retweeted: {username}")
                else:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'Retweet failed'
                    })
                
                time.sleep(2)
                    
            except Exception as e:
                results.append({
                    'username': username,
                    'status': 'failed',
                    'error': str(e)
                })
        
        return {
            "success": True,
            "retweeted_count": successful_retweets,
            "failed_count": len(request.attendees) - successful_retweets,
            "results": results
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/like-posts")
async def like_posts(
    request: TwitterActionRequest, 
    db: Session = Depends(get_db)
):
    """FIXED: Like posts using v2 API with database logging"""
    try:
        print(f"❤️  LIKING {len(request.attendees)} posts")
        
        # Update analytics
        crud.update_analytics(db, "twitter_action")
        
        twitter_client = TwitterClient()
        
        if not twitter_client.is_operational():
            return {"success": False, "error": "Twitter client not operational"}

        results = []
        successful_likes = 0
        
        for attendee in request.attendees:
            try:
                username = attendee.get('username', '')
                post_link = attendee.get('post_link', '')
                
                if not post_link:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'No post link available'
                    })
                    continue
                
                tweet_id = extract_tweet_id(post_link)
                if not tweet_id:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'Could not extract tweet ID from link'
                    })
                    continue
                
                print(f"   ❤️  Liking {username}'s tweet: {tweet_id}")
                
                # FIXED: Use client_v2.like_tweet
                like_result = twitter_client.like_tweet(tweet_id)
                
                # Log action to database
                action_data = schemas.UserActionCreate(
                    action_type="like",
                    target_username=username,
                    target_tweet_id=tweet_id,
                    status="success" if like_result else "failed",
                    message=f"Liked post from {username}",
                    additional_data={"post_link": post_link}
                )
                crud.create_user_action(db, action_data)
                
                if like_result:
                    successful_likes += 1
                    results.append({
                        'username': username,
                        'status': 'liked',
                        'tweet_id': tweet_id,
                        'message': f'Successfully liked post from {username}'
                    })
                    print(f"   ✅ Liked: {username}")
                else:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'Like failed'
                    })
                
                time.sleep(2)
                    
            except Exception as e:
                results.append({
                    'username': username,
                    'status': 'failed',
                    'error': str(e)
                })
        
        return {
            "success": True,
            "liked_count": successful_likes,
            "failed_count": len(request.attendees) - successful_likes,
            "results": results
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/post-comments")
async def post_comments(
    request: TwitterActionRequest, 
    db: Session = Depends(get_db)
):
    """FIXED: Post comments using v2 API with database logging"""
    try:
        print(f"💬 POSTING COMMENTS on {len(request.attendees)} posts")
        
        # Update analytics
        crud.update_analytics(db, "twitter_action")
        
        twitter_client = TwitterClient()
        
        if not twitter_client.is_operational():
            return {"success": False, "error": "Twitter client not operational"}

        results = []
        successful_posts = 0
        
        for attendee in request.attendees:
            try:
                username = attendee.get('username', '')
                post_link = attendee.get('post_link', '')
                custom_message = request.message or "Great post! 👍"
                
                if not post_link:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'No post link available'
                    })
                    continue
                
                # Extract tweet ID from post link
                tweet_id = extract_tweet_id(post_link)
                if not tweet_id:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'Could not extract tweet ID from link'
                    })
                    continue
                
                # Create comment text
                clean_username = username.replace('@', '')
                comment_text = f"@{clean_username} {custom_message}"
                
                # Validate comment length (Twitter limit is 280 characters)
                if len(comment_text) > 280:
                    # Truncate the custom message if needed, keeping the @username
                    max_custom_len = 280 - len(f"@{clean_username} ") - 3  # -3 for "..."
                    if max_custom_len > 0:
                        truncated_message = custom_message[:max_custom_len] + "..."
                        comment_text = f"@{clean_username} {truncated_message}"
                    else:
                        # Username is too long, just use username
                        comment_text = f"@{clean_username}"
                
                print(f"   💬 Commenting on {username}'s tweet: {tweet_id}")
                print(f"   📝 Comment text ({len(comment_text)} chars): {comment_text[:100]}...")
                
                # FIXED: Use client_v2.post_tweet instead of client.api
                result = twitter_client.post_tweet(comment_text, tweet_id)
                
                # Validate result exists and has expected structure
                if not result:
                    error_msg = "No response from Twitter API"
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': error_msg
                    })
                    print(f"   ❌ Comment failed for {username}: {error_msg}")
                    # Log failed action
                    action_data = schemas.UserActionCreate(
                        action_type="comment",
                        target_username=username,
                        target_tweet_id=tweet_id,
                        status="failed",
                        message=f"Failed to comment on post from {username}",
                        additional_data={"post_link": post_link, "comment_text": comment_text, "error": error_msg}
                    )
                    crud.create_user_action(db, action_data)
                    continue
                
                # Log action to database
                action_data = schemas.UserActionCreate(
                    action_type="comment",
                    target_username=username,
                    target_tweet_id=tweet_id,
                    status="success" if result.get('success') else "failed",
                    message=f"Commented on post from {username}",
                    additional_data={"post_link": post_link, "comment_text": comment_text, "comment_id": result.get('tweet_id')}
                )
                crud.create_user_action(db, action_data)
                
                # Check result safely
                if result.get('success'):
                    successful_posts += 1
                    results.append({
                        'username': username,
                        'status': 'commented',
                        'tweet_id': tweet_id,
                        'comment_id': result.get('tweet_id'),
                        'comment_text': comment_text,
                        'message': f'Successfully commented on post from {username}'
                    })
                    print(f"   ✅ Comment posted to {username}")
                else:
                    error_msg = result.get('error', 'Unknown error')
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': error_msg
                    })
                    print(f"   ❌ Comment failed for {username}: {error_msg}")
                
                time.sleep(3)  # Rate limiting
                
            except Exception as e:
                results.append({
                    'username': username,
                    'status': 'failed',
                    'error': str(e)
                })
                print(f"   ❌ Comment failed for {username}: {e}")
        
        # Analyze failures to provide helpful feedback
        failed_403 = sum(1 for r in results if r.get('status') == 'failed' and '403' in str(r.get('error', '')))
        failed_other = len(request.attendees) - successful_posts - failed_403
        
        response_data = {
            "success": True,
            "commented_count": successful_posts,
            "failed_count": len(request.attendees) - successful_posts,
            "results": results
        }
        
        # Add helpful message if many 403 errors
        if failed_403 > 0:
            response_data["warning"] = (
                f"{failed_403} comment(s) failed due to Twitter restrictions (403 Forbidden). "
                "This is common with automated accounts. Twitter may block automated replies. "
                "Consider using Direct Messages instead, or manually posting comments."
            )
            print(f"⚠️ {failed_403} comments blocked by Twitter (403 Forbidden)")
        
        return response_data
        
    except Exception as e:
        print(f"❌ Comment endpoint error: {e}")
        return {"success": False, "error": str(e)}

@app.post("/api/post-quote-tweets")
async def post_quote_tweets(
    request: TwitterActionRequest, 
    db: Session = Depends(get_db)
):
    """Post quote tweets using OAuth 1.1 with database logging"""
    try:
        print(f"🔁 POSTING QUOTE TWEETS for {len(request.attendees)} posts")
        
        # Update analytics
        crud.update_analytics(db, "twitter_action")
        
        twitter_client = TwitterClient()
        
        if not twitter_client.api_v1:
            return {
                "success": False,
                "error": "Twitter OAuth 1.1 not configured for quote tweets"
            }
        
        results = []
        successful_quotes = 0
        
        for attendee in request.attendees:
            try:
                username = attendee.get('username', '')
                post_link = attendee.get('post_link', '')
                custom_message = request.message or "Check this out! 👀"
                
                if not post_link:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'No post link available'
                    })
                    continue
                
                # Extract tweet ID from post link
                tweet_id = extract_tweet_id(post_link)
                if not tweet_id:
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': 'Could not extract tweet ID from link'
                    })
                    continue
                
                # Create quote tweet text
                clean_username = username.replace('@', '')
                quote_text = f"{custom_message}\n\n🔁 Via @{clean_username}"
                
                # POST QUOTE TWEET USING OAUTH 1.1
                print(f"   🔁 Creating quote tweet for {username}'s tweet: {tweet_id}")
                
                # For OAuth 1.1, we use retweet with comment (quote tweet)
                tweet = twitter_client.api_v1.update_status(
                    status=quote_text
                )
                
                # Log action to database
                action_data = schemas.UserActionCreate(
                    action_type="quote_tweet",
                    target_username=username,
                    target_tweet_id=tweet_id,
                    status="success",
                    message=f"Quoted post from {username}",
                    additional_data={"post_link": post_link, "quote_text": quote_text, "quote_tweet_id": tweet.id}
                )
                crud.create_user_action(db, action_data)
                
                successful_quotes += 1
                results.append({
                    'username': username,
                    'status': 'quoted',
                    'original_tweet_id': tweet_id,
                    'quote_tweet_id': tweet.id,
                    'quote_text': quote_text,
                    'message': f'Successfully quoted post from {username}'
                })
                print(f"   ✅ Quote tweet posted for {username}")
                
                # Add delay to avoid rate limits
                time.sleep(3)
                    
            except Exception as e:
                results.append({
                    'username': username,
                    'status': 'failed',
                    'error': str(e)
                })
                print(f"   ❌ Quote tweet failed for {username}: {e}")
        
        return {
            "success": True,
            "quoted_count": successful_quotes,
            "failed_count": len(request.attendees) - successful_quotes,
            "total_attempted": len(request.attendees),
            "results": results
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/send-messages")
async def send_messages(
    request: TwitterActionRequest, 
    db: Session = Depends(get_db)
):
    """Send direct messages to selected users"""
    try:
        print(f"💬 SENDING MESSAGES to {len(request.attendees)} users")
        
        # Update analytics
        crud.update_analytics(db, "twitter_action")
        
        from services.unified_messaging import UnifiedMessagingService
        
        messaging_service = UnifiedMessagingService()
        
        if not messaging_service.is_platform_ready("twitter"):
            return {
                "success": False,
                "error": "Twitter client not operational"
            }
        
        if not request.message or not request.message.strip():
            return {
                "success": False,
                "error": "Message is required"
            }
        
        results = []
        successful_messages = 0
        
        for attendee in request.attendees:
            try:
                username = attendee.get('username', '').replace('@', '')
                user_id = attendee.get('user_id') or attendee.get('id')
                
                # If user_id is missing, try to look it up from username
                if not user_id:
                    print(f"   🔍 Looking up user_id for @{username}")
                    user_id = await lookup_user_id_from_username(username)
                    if not user_id:
                        results.append({
                            'username': username,
                            'status': 'failed',
                            'error': 'Could not find user_id for username. User may not exist or account may be private.'
                        })
                        continue
                
                print(f"   💬 Sending DM to @{username} (user_id: {user_id})")
                
                # Prepare target dict for messaging service
                target = {
                    'username': username,
                    'user_id': user_id
                }
                
                # Send message using unified messaging service
                result = messaging_service.send_message(
                    platform='twitter',
                    target=target,
                    message=request.message
                )
                
                # Log action to database
                action_data = schemas.UserActionCreate(
                    action_type="direct_message",
                    target_username=username,
                    target_tweet_id="",  # DMs don't have tweet IDs
                    status="success" if result.get('success') else "failed",
                    message=f"Sent DM to {username}",
                    additional_data={"message": request.message[:100], "user_id": user_id}
                )
                crud.create_user_action(db, action_data)
                
                if result.get('success'):
                    successful_messages += 1
                    results.append({
                        'username': username,
                        'status': 'sent',
                        'message': f'Successfully sent DM to {username}'
                    })
                    print(f"   ✅ DM sent to @{username}")
                else:
                    error_msg = result.get('error', 'Unknown error')
                    results.append({
                        'username': username,
                        'status': 'failed',
                        'error': error_msg
                    })
                    print(f"   ❌ DM failed for @{username}: {error_msg}")
                
                # Rate limiting - wait between messages
                time.sleep(2)
                    
            except Exception as e:
                results.append({
                    'username': username,
                    'status': 'failed',
                    'error': str(e)
                })
                print(f"   ❌ DM failed for {username}: {e}")
        
        return {
            "success": True,
            "sent_count": successful_messages,
            "failed_count": len(request.attendees) - successful_messages,
            "results": results
        }
        
    except Exception as e:
        print(f"❌ Send messages endpoint error: {e}")
        return {"success": False, "error": str(e)}

async def lookup_user_id_from_username(username: str) -> Optional[str]:
    """Look up Twitter user_id from username"""
    try:
        from services.twitter_client import TwitterClient
        
        twitter_client = TwitterClient()
        
        if not twitter_client.is_operational():
            return None
        
        # Clean username (remove @ if present)
        clean_username = username.replace('@', '').strip()
        
        # Use v2 API to get user by username
        try:
            if twitter_client.client_v2 is None:
                return None
            user = twitter_client.client_v2.get_user(username=clean_username)
            if user and user.data:
                return str(user.data.id)
        except Exception as e:
            print(f"   ⚠️ Could not look up user_id for @{clean_username}: {e}")
            return None
        
        return None
    except Exception as e:
        print(f"   ⚠️ User lookup error: {e}")
        return None

@app.post("/api/test-single-comment")
async def test_single_comment():
    """Test endpoint for posting a single comment"""
    try:
        from services.twitter_client import TwitterClient
        
        twitter_client = TwitterClient()
        
        if not twitter_client.api_v1:
            return {"success": False, "error": "OAuth 1.1 not available"}
        
        # Test with a known tweet ID
        test_tweet_id = "1879999999999999999"  # Replace with actual tweet ID
        test_username = "testuser"
        comment_text = f"@{test_username} 👋 Test comment from Event Intelligence Platform! 🎉"
        
        print(f"🧪 Testing comment on tweet: {test_tweet_id}")
        
        tweet = twitter_client.api_v1.update_status(
            status=comment_text,
            in_reply_to_status_id=test_tweet_id,
            auto_populate_reply_metadata=True
        )
        
        return {
            "success": True,
            "message": "Test comment posted successfully",
            "tweet_id": tweet.id,
            "original_tweet_id": test_tweet_id
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

# Get analytics endpoint
@app.get("/api/analytics")
async def get_analytics(db: Session = Depends(get_db)):
    """Get platform analytics"""
    try:
        # Get today's analytics
        today = datetime.utcnow().date()
        
        from database import models
        analytics = db.query(models.Analytics).filter(
            func.date(models.Analytics.date) == today
        ).first()
        
        if analytics:
            return {
                "success": True,
                "analytics": {
                    "event_searches": analytics.event_searches,
                    "attendee_searches": analytics.attendee_searches,
                    "twitter_actions": analytics.twitter_actions,
                    "api_calls_saved": analytics.api_calls_saved,
                    "active_users": analytics.active_users
                }
            }
        else:
            return {
                "success": True,
                "analytics": {
                    "event_searches": 0,
                    "attendee_searches": 0,
                    "twitter_actions": 0,
                    "api_calls_saved": 0,
                    "active_users": 0
                }
            }
    except Exception as e:
        return {"success": False, "error": str(e)}

# Get search history endpoint
@app.get("/api/search-history")
async def get_search_history(
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """Get recent search history"""
    try:
        from database import models
        from sqlalchemy import desc
        
        history = db.query(models.SearchHistory).order_by(
            desc(models.SearchHistory.created_at)
        ).limit(limit).all()
        
        return {
            "success": True,
            "history": [
                {
                    "id": h.id,
                    "search_type": h.search_type,
                    "query": h.query,
                    "results_count": h.results_count,
                    "created_at": h.created_at.isoformat()
                }
                for h in history
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

def extract_tweet_id(post_link: str) -> Optional[str]:
    """Extract tweet ID from Twitter post link"""
    try:
        patterns = [
            r'status/(\d+)',
            r'twitter\.com/\w+/status/(\d+)',
            r'twitter\.com/status/(\d+)',
            r'x\.com/\w+/status/(\d+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, post_link)
            if match:
                return match.group(1)
        
        return None
    except Exception:
        return None

# Excel Export Endpoints
@app.post("/api/export-events")
async def export_events(events_data: Dict[str, Any]):
    """Export events to Excel file"""
    try:
        events = events_data.get("events", [])
        if not events:
            raise HTTPException(status_code=400, detail="No events to export")
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=500, detail="Failed to create worksheet")
        ws.title = "Events"
        
        # Define headers
        headers = ["Event Name", "Date", "Time", "Venue", "Location", "Category", "Source", "Hype Score", "Source URL"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data rows
        for event in events:
            # Parse date and time
            date_str = event.get("exact_date", "")
            date_part = ""
            time_part = ""
            if date_str:
                try:
                    # Try to split date and time
                    if " " in date_str:
                        parts = date_str.split(" ", 1)
                        date_part = parts[0]
                        if len(parts) > 1:
                            time_part = parts[1]
                    else:
                        date_part = date_str
                except:
                    date_part = date_str
            
            row = [
                event.get("event_name", ""),
                date_part,
                time_part,
                event.get("exact_venue", ""),
                event.get("location", ""),
                event.get("category", ""),
                event.get("source", ""),
                event.get("hype_score", 0.0),
                event.get("source_url", "")
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        column_widths = {
            "A": 40,  # Event Name
            "B": 15,  # Date
            "C": 15,  # Time
            "D": 30,  # Venue
            "E": 20,  # Location
            "F": 15,  # Category
            "G": 12,  # Source
            "H": 12,  # Hype Score
            "I": 50   # Source URL
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"events_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ ERROR in export_events: {e}")
        print(error_trace)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@app.post("/api/export-attendees")
async def export_attendees(attendees_data: Dict[str, Any]):
    """Export attendees to Excel file"""
    try:
        attendees = attendees_data.get("attendees", [])
        if not attendees:
            raise HTTPException(status_code=400, detail="No attendees to export")
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=500, detail="Failed to create worksheet")
        ws.title = "Attendees"
        
        # Define headers
        headers = ["Username", "Source", "Engagement Type", "Post Content", "Post Date", "Location", "Followers", "Verified", "Post Link"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data rows
        for attendee in attendees:
            row = [
                attendee.get("username", ""),
                attendee.get("source", ""),
                attendee.get("engagement_type", ""),
                attendee.get("post_content", ""),
                attendee.get("post_date", ""),
                attendee.get("location", ""),
                attendee.get("followers_count", 0),
                "Yes" if attendee.get("verified", False) else "No",
                attendee.get("post_link", "")
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        column_widths = {
            "A": 20,  # Username
            "B": 12,  # Source
            "C": 18,  # Engagement Type
            "D": 60,  # Post Content
            "E": 15,  # Post Date
            "F": 20,  # Location
            "G": 12,  # Followers
            "H": 10,  # Verified
            "I": 50   # Post Link
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendees_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ ERROR in export_attendees: {e}")
        print(error_trace)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

# Serve frontend
# -----------------------------
# FRONTEND SETUP (FIXED)
# -----------------------------

# Get the frontend directory path (parent directory from Backend)
# Use __file__ to get the absolute path of this file, then go up one level
try:
    BASE_DIR = Path(__file__).resolve().parent.parent
except NameError:
    # Fallback if __file__ is not available
    BASE_DIR = Path.cwd().parent
FRONTEND_DIR = BASE_DIR / "frontend"

# 1) Serve static files (CSS, JS, images)
app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")
 
# 2) Serve index.html for root (with no-cache headers)
@app.get("/")
async def serve_frontend():
    index_path = FRONTEND_DIR / "index.html"
    response = FileResponse(str(index_path))
    # Add no-cache headers to prevent caching
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

if __name__ == "__main__":
    print("🚀 Event Intelligence Platform Starting...")
    print("📡 API: http://localhost:8000")
    print("🎯 POLICY: FIXED - OAuth 1.1 for all Twitter actions")
    print("🐦 TWITTER ACTIONS: Retweet, Like, Comment, Quote Tweet")
    print("💾 DATABASE: PostgreSQL with caching and analytics")
    print("💡 Test auth status: http://localhost:8000/api/auth-status")
    uvicorn.run(app, host="0.0.0.0", port=8000)